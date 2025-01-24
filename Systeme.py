import psutil
import time
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import socket
import re


# Create an Excel file and initialize the sheet
def create_excel_file(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resource Usage"

    # Add column headers
    headers = ["Date", "RAM Usage (GB)", "CPU Usage (%)", "Network Usage (MB/s)", "Number of HTTP/HTTPS Connections"]
    ws.append(headers)

    # Format the headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = Border(bottom=Side(style="thin"))

    # Adjust the column widths
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 25

    wb.save(file_name)


# Function to get RAM usage in GB
def get_ram_usage_in_gb():
    ram = psutil.virtual_memory()
    return round(ram.used / (1024 ** 3), 2)


# Function to get CPU usage in %
def get_cpu_usage_in_percent():
    return psutil.cpu_percent(interval=1)


# Function to get network usage in MB/s
def get_network_usage_in_mbps():
    net_io = psutil.net_io_counters()
    bytes_sent = net_io.bytes_sent
    bytes_recv = net_io.bytes_recv

    time.sleep(1)

    net_io = psutil.net_io_counters()
    sent_per_sec = (net_io.bytes_sent - bytes_sent) / (1024 * 1024)
    recv_per_sec = (net_io.bytes_recv - bytes_recv) / (1024 * 1024)

    return round(sent_per_sec + recv_per_sec, 2)


# Function to get the number of HTTP/HTTPS connections
def get_http_connections_count():
    http_count = 0
    for conn in psutil.net_connections(kind='inet'):
        if conn.status == 'ESTABLISHED' and conn.raddr and conn.raddr.port in (80, 443):
            http_count += 1
    return http_count


# Function to get the domain name from an IP address
def get_domain_from_ip(ip):
    try:
        domain = socket.gethostbyaddr(ip)
        domain_name = domain[0]

        if re.match(r"^(?:\d{1,3}\.){3}\d{1,3}$", domain_name) or "ip-" in domain_name:
            return None

        if domain_name and "." in domain_name:
            return domain_name
        else:
            return None
    except (socket.herror, socket.gaierror):
        return None


# Function to collect data on HTTP/HTTPS requests
def get_domain_usage_data():
    domain_data = {}

    for conn in psutil.net_connections(kind='inet'):
        if conn.status == 'ESTABLISHED' and conn.raddr and conn.raddr.port in (80, 443):
            ip = conn.raddr.ip
            domain = get_domain_from_ip(ip)

            if domain:
                if domain not in domain_data:
                    domain_data[domain] = {'request_count': 0}
                domain_data[domain]['request_count'] += 1

    return domain_data


# Function to apply conditional colors
def apply_colors(ws, col_idx):
    values = [ws.cell(row=row, column=col_idx).value for row in range(2, ws.max_row + 1)]
    average_value = sum(values) / len(values) if values else 0

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value > average_value:
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif cell.value < average_value:
            cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")


# Function to add individual charts
def add_individual_charts(ws):
    chart_positions = ["G2", "G20", "G38", "G56"]
    titles = ["RAM (GB)", "CPU (%)", "Network (MB/s)", "HTTP/HTTPS Connections"]

    for i, col_idx in enumerate(range(2, 6), start=0):
        chart = LineChart()
        chart.title = titles[i]
        chart.y_axis.title = titles[i]
        chart.x_axis.title = "Time"

        data = Reference(ws, min_col=col_idx, min_row=1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws.add_chart(chart, chart_positions[i])


# Function to apply black borders around all cells
def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000")
            )


# Function to save deleted data
def save_deleted_data(deleted_data):
    file_name = "OLD_DATA.xlsx"
    if not os.path.exists(file_name):
        create_excel_file(file_name)

    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    for row in deleted_data:
        ws.append(row)

    for col_idx in range(2, 6):
        apply_colors(ws, col_idx)

    apply_borders(ws)
    add_individual_charts(ws)
    wb.save(file_name)


# Function to save domain usage data with hyperlinks
# Function to save domain usage data with hyperlinks
def save_domain_usage_data(domain_data, file_name="Domain_Usage_Report.xlsx"):
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Domain Usage"

        # Add headers
        headers = ["Domain Name", "Request Count", "Network Usage Percentage"]
        ws.append(headers)

        # Calculate total requests
        total_requests = sum(data['request_count'] for data in domain_data.values())

        # Fill in domain data
        for domain, data in domain_data.items():
            if domain:
                percentage_usage = (data['request_count'] / total_requests) * 100 if total_requests > 0 else 0
                ws.append([domain, data['request_count'], round(percentage_usage, 2)])
                ws.cell(row=ws.max_row, column=1).hyperlink = f"http://{domain}"
                ws.cell(row=ws.max_row, column=1).font = Font(underline="single", color="0000FF")

        # Format headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = Border(bottom=Side(style="thin"))

        # Adjust column widths
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 30

        # Apply colors and borders
        apply_colors(ws, 2)
        apply_borders(ws)

        # Add a chart
        chart = LineChart()
        chart.title = "Domain Usage"
        chart.y_axis.title = "Request Count"
        chart.x_axis.title = "Domain Name"

        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws.add_chart(chart, "E5")

        # Save the workbook
        wb.save(file_name)
        print(f"Domain usage data saved successfully to '{file_name}'.")

    except PermissionError:
        print(f"Permission denied: Unable to save the file '{file_name}'. Please close the file if it is open.")
    except Exception as e:
        print(f"An error occurred while saving domain usage data: {e}")



# Main function to log system usage
def log_system_usage(num):
    main_file = "System_Report.xlsx"
    if not os.path.exists(main_file):
        create_excel_file(main_file)

    wb = openpyxl.load_workbook(main_file)
    ws = wb.active

    max_data_rows = 40
    deleted_data = []

    while ws.max_row > max_data_rows:
        deleted_row = [ws.cell(row=2, column=col).value for col in range(1, 6)]
        deleted_data.append(deleted_row)
        ws.delete_rows(2)

    save_deleted_data(deleted_data)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ram_usage = get_ram_usage_in_gb()
    cpu_usage = get_cpu_usage_in_percent()
    network_usage = get_network_usage_in_mbps()
    http_count = get_http_connections_count()

    domain_data = get_domain_usage_data()

    ws.append([timestamp, ram_usage, cpu_usage, network_usage, http_count])

    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(2, 6):
        apply_colors(ws, col_idx)

    apply_borders(ws)
    add_individual_charts(ws)
    wb.save(main_file)

    save_domain_usage_data(domain_data)
    print(f"{num} - {timestamp} - RAM: {ram_usage} GB, CPU: {cpu_usage}%, Network: {network_usage} MB/s, Connections: {http_count}")


# Periodic execution
if __name__ == "__main__":
    num = 0
    while True:
        num += 1
        log_system_usage(num)
        time.sleep(1)
